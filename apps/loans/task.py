from celery import shared_task
from openpyxl import load_workbook
from apps.customers.models import Customers
from apps.loans.models import Loans



@shared_task
def seed_data(customer_file_path, loan_file_path):
    customer_wb = load_workbook(customer_file_path)
    customer_sheet = customer_wb.active

    loan_wb = load_workbook(loan_file_path)
    loan_sheet = loan_wb.active

    customer_map = {}

    for row in customer_sheet.iter_rows(min_row=2, values_only=True):
        customer_id, first_name, last_name, age, phone, monthly_salary, approved_limit = row
        if customer_id is not None:
            customer = Customers.objects.create(
                first_name=first_name,
                last_name=last_name,
                age=age,
                phone_number=phone,
                monthly_salary=monthly_salary,
                approved_limit=approved_limit
            )
            customer_map[customer_id] = customer

    duplicate=set()
    for row in loan_sheet.iter_rows(min_row=2, values_only=True):
        customer_id, loan_id, amount, tenure, rate, monthly_payment, emi_on_time, approval_date, end_date = row

        if(loan_id in duplicate):
            print(f"Skipping duplicate loan entry {loan_id}")
            continue

        duplicate.add(loan_id)
        
        if customer_id is not None and customer_id in customer_map:
            customer = customer_map[customer_id]
            Loans.objects.create(
                customer=customer,
                tenure=tenure,
                interest_rate=rate,
                loan_amount=amount,
                monthly_payment=monthly_payment,
                emi_paid_on_time=emi_on_time,
                date_of_approval=approval_date,
                end_date=end_date
            )
        else:
            print(f"Customer {customer_id} not found")
           
    customers = Customers.objects.all()

    for customer in customers:
        loans = Loans.objects.filter(customer=customer)
        total_debt = sum(max(loan.tenure - loan.emi_paid_on_time, 0) * loan.monthly_payment for loan in loans)
        customer.current_debt = total_debt
        customer.save()


    print("Data seeded successfully")