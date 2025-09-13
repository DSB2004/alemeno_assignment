from openpyxl import load_workbook
import os
import django
from openpyxl import load_workbook
import sys


sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")  
django.setup()

from apps.customers.models import Customers

def add_customers():
    wb = load_workbook("./assets/customer_data.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        customer_id, first_name,last_name, age, phone ,monthly_salary,approved_limit= row
        print(customer_id,first_name,last_name, age, phone ,monthly_salary,approved_limit)
        if customer_id is not None:
            Customers.objects.create(
                customer_id=customer_id,
                first_name=first_name,
                last_name=last_name,
                monthly_salary=monthly_salary,
                approved_limit=approved_limit,
                phone_number=phone,
                age=age
                
            )

    print("Customers added successfully")


if __name__=="__main__":
    add_customers()