from openpyxl import load_workbook
import os
import django
from openpyxl import load_workbook
import sys


sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")  
django.setup()

from apps.loans.models import Loans
from apps.customers.models import Customers
def add_loans():
    wb = load_workbook("./assets/loan_data.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        customer_id,loan_id, amount,tenure,rate,monthly_payment,emi_on_time,approval_date,end_date= row
        print(customer_id,loan_id, amount,tenure,rate,monthly_payment,emi_on_time,approval_date,end_date)
        if customer_id is not None:
            try:
                Loans.objects.create(
                    loan_id=loan_id,
                    customer_id=customer_id,
                    tenure=tenure,
                    interest_rate=rate,
                    monthly_payment=monthly_payment,
                    emi_paid_on_time=emi_on_time,
                    date_of_approval=approval_date,
                    end_date=end_date
                )
            except Customers.DoesNotExist:
                print(f"Customer {customer_id} not found")
            except Exception as e:
                print(e)
                print(f"Error adding loan {loan_id}")

    print("Customers added successfully")


if __name__=="__main__":
    add_loans()