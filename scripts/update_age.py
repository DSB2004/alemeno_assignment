from openpyxl import load_workbook
import os
import django
from openpyxl import load_workbook
import sys


sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")  
django.setup()

from apps.customers.models import Customers

def update_age():
    wb = load_workbook("./assets/customer_data.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            customer_id, first_name,last_name, age, phone ,monthly_salary,approved_limit= row
            if(customer_id is not None):
                customer=Customers.objects.get(id=customer_id)
                customer.age=int(age)
                customer.save()
        except Customers.DoesNotExist:
            print("No exist",customer_id)
        

    print("Customers age updated successfully")


if __name__=="__main__":
    update_age()